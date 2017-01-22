# -*- coding: utf-8 -*-
"""
Created on Tue Jan  3 09:37:52 2017

@author: K. Biß

Aim is the projection of RE according to EEG "Zubaupfade" and retirement of old facilities
Also checking the corresponding TWh with goals of the goverment
"""


import openpyxl,  numpy as np
from openpyxl.cell import get_column_letter

iStart  = 1990
iEnd    = 2100
hpyear  = 8760
strWorkbook = "EE_installierteLeistung.xlsx"
wb=openpyxl.load_workbook(strWorkbook)

##################Sterbelinien hinzufügen#######################
ws_phaseout = wb.get_sheet_by_name('decay')
standardDecay = [ [ loc.value for loc in rowloc] for rowloc in ws_phaseout['B43':'AC43']]
'''for rowloc in ws_phaseout['B43':'AC43']:
    for location in rowloc:
        standardDecay.append(location.value)
'''
fastDecay = [[cell.value for cell in rowloc] for rowloc in ws_phaseout['B44':'L44'] ]        
fourDecay =[100, 50, 25, 0]
listDecay = [standardDecay, fastDecay, fourDecay]        
################################################################

def capMatrix(Anlagentypen, DecayList, minJahr, maxJahr, ecolife, RE_cap):
    iStart = minJahr
    iEnd=maxJahr
    dif = iEnd-iStart
    Anlagennamen = Anlagentypen
    mStd = np.zeros( (len(Anlagennamen),dif+1, dif+1), dtype=np.double) 
    for i in range(0, len(Anlagennamen)):
        for build in range(0,dif+1):
            for adj in range(0,dif+1):
                if adj == build:
                    mStd[i,build,build] = RE_cap[Anlagennamen[i]][iStart+build]['cap_year'] 
                elif adj > build:
                    dlife = adj-build
                    if dlife < ecolife[Anlagennamen[i]]:
                        mStd[i,build,adj] = mStd[i,build,build] 
                    else:
                        try:
                            mStd[i,build,adj] = mStd[i,build,build] * DecayList[dlife-ecolife[Anlagennamen[i]]]/100
                        except IndexError: # dann ist Decayliste überschritten
                            mStd[i,build,adj] = 0                         
    return mStd

# wenn numnpy verstanden, funktion überflüssig
def sumCapMatrix(Anlagentypen,  minJahr, maxJahr, capMatrixData):
        iStart = minJahr
        iEnd=maxJahr
        dif = iEnd-iStart
        sumfirst =0
        resultArr = {}
        for i in range(0,len(Anlagennamen)):
            for j in range(0,dif+1): 
                for k in range(0,dif+1):
                    sumfirst +=  capMatrixData[i,k,j]
                    resultArr.setdefault(Anlagennamen[i],{})
                    resultArr[Anlagennamen[i]].setdefault( iStart+j, {'cap_total':0} )             
                    resultArr[Anlagennamen[i]][iStart+j]['cap_total'] = sumfirst
                sumfirst =0
        return     resultArr

def sumCapMatrixShort(Anlagentypen, capMatrixData):
        resultArr = {}
        for i in range(0,len(Anlagentypen)):
            resultArr.setdefault(Anlagentypen[i],[])
            resultArr[Anlagentypen[i]]= capMatrixData[i,:,:].sum(axis=0)
        return     resultArr
#Add on Check der Regierungsziele        
def sumPowerMatrix(capMatrixData, minJahr, maxJahr, Anlagentyp):
    iStart = minJahr
    iEnd   = maxJahr
    dif = iEnd-iStart
    hpyear = 8760
    sumPower = np.full( (dif+1), 0, dtype = np.double)
    for k in capMatrixData:
        CtAfactor = available(k) #  weil dictornary nicht verstanden
        for j in range(0,dif +1):
            sumPower[j] += capMatrixData[k][iStart + j]['cap_total'] * CtAfactor * hpyear / 1000000.0 # MW --> TWh
    return sumPower
            
            

'''
resultStdCap = sumCapMatrix(Anlagennamen,  iStart, iEnd, mStandard)
resultList= [resultStdCap, resultfastCap, result4yCap, resultnoneCap]
resultListName= ['mStandard', 'mfast', 'm4Years', 'mPolitics']

test = sumPowerMatrix(resultStdCap, iStart, iEnd, Anlagennamen)'''
def available(Anlagentyp):
    factors =  {'Water': 0.63, 'WEAon': 0.23, 'WEAoff':0.56, 'Bio':0.70, 'PV':0.13, 'Geo':0}
    factor = factors[Anlagentyp]
    return factor

#import installierte Leistung
wsPast = wb.get_sheet_by_name('historic')

Anlagentypen = [[ cell.value for cell in Header ] for Header in wsPast['B3':'G3']]
Anlagennamen2 = Anlagentypen[0] 
       
RE_cap = {}
for data in range(4, wsPast.max_row -2): # Spalten
    intYear = wsPast[get_column_letter(1)+str(data)].value
    for col in range(2, wsPast.max_column+1): 
        tempType = Anlagennamen[col-2] # wegen Liste und Excel 0 und 1
        RE_cap.setdefault(tempType,{})
        RE_cap[tempType].setdefault(intYear, {'cap_total':0, 'cap_year':0})                   
        RE_cap[tempType][intYear]['cap_total'] = wsPast.cell(row=data, column=col).value

#Add Zubaupfad
zubau = wb.get_sheet_by_name('zubau')    
for Anlagentyp in range(2, zubau.max_row+1):
    for col in range(2, zubau.max_column+1): 
        intYear = zubau[get_column_letter(col)+'1'].value
        strAnlage = zubau.cell(row=Anlagentyp, column=1).value
        RE_cap[strAnlage].setdefault(intYear, {'cap_total':0, 'cap_year':0})
        RE_cap[strAnlage][intYear]['cap_year']= zubau.cell(row=Anlagentyp, column=col).value
        
#Erzeuge fehlende Daten
#2030+ 
for k  in RE_cap:
    for i in range(iStart,iEnd+1):
        if i == iStart:
            RE_cap[k][i]['cap_year'] = RE_cap[k][i]['cap_total']
        else:
            try:
                if RE_cap[k][i]['cap_total']>0 :
                    RE_cap[k][i]['cap_year'] = RE_cap[k][i]['cap_total']-RE_cap[k][i-1]['cap_total']
                    if RE_cap[k][i]['cap_year'] < 0:
                        RE_cap[k][i]['cap_year'] = 0 # Bei Wasser der Fall oder wenn es wirklich zu Nettoabnahme kommt
                        # if RE_cap[k][i]['cap_total']-RE_cap[k][i-1]['cap_total'] <0:
                        # print ('Fehler in historischen Werten bei' + k + str(i))
                else:
                    try:
                        RE_cap[k][i]['cap_total'] = RE_cap[k][i-1]['cap_total'] + RE_cap[k][i]['cap_year']
                    except TypeError: # der Fall, wenn kein Zubau angegeben
                        RE_cap[k][i]['cap_total'] = RE_cap[k][i-1]['cap_total']
                        RE_cap[k][i]['cap_year'] = 0
            except KeyError: 
                RE_cap[k].setdefault(i, {'cap_total':0, 'cap_year':0}) 
                RE_cap[k][i]['cap_year'] = RE_cap[k][i-1]['cap_year'] # Forschreibung mit letztem Ausbau (Annahme!)
                RE_cap[k][i]['cap_total'] = RE_cap[k][i-1]['cap_total'] + RE_cap[k][i]['cap_year']
                
            
#Sterbelinien hinzufügen
ws_phaseout = wb.get_sheet_by_name('decay')
standardDecay = []
fastDecay = []
fourDecay =[100, 50, 25, 0]
noneDecay = np.full( (100), 100, dtype = np.double)
ListDecay = [standardDecay, fastDecay, fourDecay, noneDecay]
ecolife = {'Water': 100, 'WEAon':25, 'WEAoff':20, 'Bio':20, 'PV':20, 'Geo':25} # Lebensdauer Wasser stimmt nicht, aber für Rechnung sollte betrag konstant bleiben, Geothermie produziert keinen strom
longlife = {'Water': 100, 'WEAon':30, 'WEAoff':25, 'Bio':30, 'PV':25, 'Geo':40}  
infilife = {'Water': 100, 'WEAon':35, 'WEAoff':30, 'Bio':40, 'PV':30, 'Geo':60} 
ListEcoLifes = [ecolife, longlife, infilife]

for rowloc in ws_phaseout['B43':'AC43']:
    for location in rowloc:
        standardDecay.append(location.value)

for rowloc in ws_phaseout['B44':'L44']:
    for location in rowloc:
        fastDecay.append(location.value)

##################################
##Datenaufnahme abgeschlossen#####
##################################


'''

##### Hard coded, flexibel wäre schöner, aber Kosmetik ######
zMatrix = np.zeros( (len(Anlagennamen),iEnd-iStart+1, iEnd-iStart+1), dtype=np.double) # first: Analgetyp; second: Baujahr third: Cap in Jahrx

mStandard = capMatrix(Anlagennamen, standardDecay, iStart, iEnd, ecolife, RE_cap)
mfast  = capMatrix(Anlagennamen, fastDecay, iStart, iEnd, ecolife, RE_cap)
m4Years  = capMatrix(Anlagennamen, fourDecay, iStart, iEnd, ecolife, RE_cap)
mPolitics = capMatrix(Anlagennamen, noneDecay, iStart, iEnd, ecolife, RE_cap )


resultStdCap = sumCapMatrix(Anlagennamen,  iStart, iEnd, mStandard)

test = sumCapMatrixShort(Anlagennamen, mStandard)

resultfastCap = sumCapMatrix(Anlagennamen,  iStart, iEnd, mfast)
result4yCap = sumCapMatrix(Anlagennamen,  iStart, iEnd, m4Years)
resultnoneCap = sumCapMatrix(Anlagennamen,  iStart, iEnd, mPolitics)

resultListCap= [resultStdCap, resultfastCap, result4yCap, resultnoneCap]
resultListName= ['mStandard', 'mfast', 'm4Years', 'mPolitics']

resStdPow = sumPowerMatrix(resultStdCap, iStart, iEnd, Anlagennamen)
resFastPow = sumPowerMatrix(resultfastCap, iStart, iEnd, Anlagennamen)
res4YPow = sumPowerMatrix(result4yCap, iStart, iEnd, Anlagennamen)
resNoPow = sumPowerMatrix(resultnoneCap, iStart, iEnd, Anlagennamen)
resultListPow= [resStdPow, resFastPow, res4YPow, resNoPow]



##Print out
wsResult = wb.get_sheet_by_name('result')


for i in range(0,len(resultListCap)):    
    rowlength = i*(len(Anlagennamen)+1)
    if i == 0: # Header
        for j in range(0,iEnd-iStart+1):
            wsResult[get_column_letter(j+2)+str(1)] = iStart + j
    # Name setzen
    for name in range(0,len(Anlagennamen)):
        wsResult[get_column_letter(1)+str(1+i*(len(Anlagennamen)+1))] = resultListName[i]
        wsResult[get_column_letter(1)+str(2+i*(len(Anlagennamen)+1)+name)] = Anlagennamen[name]
        for year in range(0,iEnd-iStart+1):
            #wsResult[get_column_letter(2+year)+str(2+rowlength+name)] = int(mfast[type[name]][iStart+year]['cap_total'])
            wsResult[get_column_letter(2+year)+str(2+rowlength+name)] = int(resultListCap[i][Anlagennamen[name]][iStart+year]['cap_total'])

for i in range(0, len(resultListName)):
    defStart = 30
    wsResult[get_column_letter(1) + str(defStart+i)]= resultListName[i]
    for v in range(0, len(resStdPow)):
        wsResult[get_column_letter(2+v) + str(defStart+i)]= resultListPow[i][v]

#TWh für 2006 bis 2012 laut BMWi
RefPow = [ 47.7  ,	62.6  ,	68.1  ,	71.5  	,79.1  ,	101.3  	,116.7  , 123.9, 136.8]
print ('Nach BMWi ab 2006: ', RefPow , ' +25 TWh Wasserkraft')           
wb.save(strWorkbook) 

'''

'''
import xlrd
from xlrd.sheet import ctype_text

wb = xlrd.open_workbook(strWorkbook)
listWs = wb.sheet_names()
ws0 = wb.sheet_by_name(listWs[1])
print ('Test: ' + ws0.name)
print ('(Column #) type:value')
'''


'''
Graphen automatisch generieren
'''


'''

print('Print results to Excel...')
resultFile = open('Ergebnis.xlsx', 'w')
resultFile.write('allData = ' + pprint.pformat(mStandard))
resultFile.close()
'''