# -*- coding: utf-8 -*-
"""
Created on Tue Jan  3 09:37:52 2017

@author: K. Biß

Aim is the projection of RE according to EEG "Zubaupfade" and retirement of old facilities
Also checking the corresponding TWh with goals of the goverment
"""


import openpyxl,  numpy as np, pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook


iStart  = 1990
iEnd    = 2100
hpyear  = 8760
strWorkbook = "EE_installierteLeistung.xlsx"
wb=openpyxl.load_workbook(strWorkbook)

##################Sterbelinien hinzufügen#######################
def readDecays():
    ws_phaseout = wb.get_sheet_by_name('decay')
    tempstandardDecay = [ [ loc.value for loc in rowloc] for rowloc in ws_phaseout['B43':'AC43']]
    '''for rowloc in ws_phaseout['B43':'AC43']:
        for location in rowloc:
            standardDecay.append(location.value)
    '''
    standardDecay = tempstandardDecay[0] ## !!!
    tempfastDecay = [[cell.value for cell in rowloc] for rowloc in ws_phaseout['B44':'L44'] ]        
    fastDecay = tempfastDecay[0] ## !!!
    fourDecay =[100, 50, 25, 0]
    noneDecay = np.full( (100), 100, dtype = np.double)
    listDecay = [standardDecay, fastDecay, fourDecay, noneDecay] 
    return listDecay

def readDecaysAsName(i):
    listDecayName = ['standardDecay', 'fastDecay', 'fourDecay', 'noneDecay']
    return listDecayName[i]
    
def readEcoLifes():
    ecolife = {'Water': 100, 'WEAon':25, 'WEAoff':20, 'Bio':20, 'PV':20, 'Geo':25} # Lebensdauer Wasser stimmt nicht, aber für Rechnung sollte betrag konstant bleiben, Geothermie produziert keinen strom
    longlife = {'Water': 100, 'WEAon':30, 'WEAoff':25, 'Bio':30, 'PV':25, 'Geo':40}  
    infilife = {'Water': 100, 'WEAon':35, 'WEAoff':30, 'Bio':40, 'PV':30, 'Geo':60} 
    listEcoLifes = [ecolife, longlife, infilife]
    return listEcoLifes
################################################################

def capMatrix(Anlagentypen, DecayList, minJahr, maxJahr, ecolife, RE_cap):
    iStart = minJahr
    iEnd=maxJahr
    dif = iEnd-iStart
    Anlagennamen = Anlagentypen
    mStd = np.zeros( (len(Anlagennamen),dif+1, dif+1), dtype=np.double) 
    for i in range(0, len(Anlagennamen)):
        for build in range(0,dif+1):
            for adj in range(0,dif+1):
                if adj == build:
                    mStd[i,build,build] = RE_cap[Anlagennamen[i]][iStart+build]['cap_year'] 
                elif adj > build:
                    dlife = adj-build
                    if dlife < ecolife[Anlagennamen[i]]:
                        mStd[i,build,adj] = mStd[i,build,build] 
                    else:
                        try:
                            mStd[i,build,adj] = mStd[i,build,build] * DecayList[dlife-ecolife[Anlagennamen[i]]]/100
                        except IndexError: # dann ist Decayliste überschritten
                            mStd[i,build,adj] = 0                         
    return mStd

def sumCapMatrix(Anlagentypen, capMatrixData):
        resultArr = {}
        for i in range(0,len(Anlagentypen)):
            resultArr.setdefault(Anlagentypen[i],[])
            resultArr[Anlagentypen[i]]= capMatrixData[i,:,:].sum(axis=0)
        return     resultArr
      
def sumPowerMatrix(capMatrixData, minJahr, maxJahr, Anlagentyp): #dict with array no specific year as key
    iStart = minJahr
    iEnd   = maxJahr
    dif = iEnd-iStart
    hpyear = 8760
    sumPower = np.full( (dif+1), 0, dtype = np.double)
    for k in capMatrixData:
        CtAfactor = available(k)
        for j in range(0,dif +1):
            sumPower[j] += capMatrixData[k][j] * CtAfactor * hpyear / 1000000.0 # MW --> TWh
    return sumPower
            
    
def available(Anlagentyp):
    factors =  {'Water': 0.63, 'WEAon': 0.23, 'WEAoff':0.56, 'Bio':0.70, 'PV':0.13, 'Geo':0}
    factor = factors[Anlagentyp]
    return factor

#import installierte Leistung
print ('Read in inital data...')
wsPast = wb.get_sheet_by_name('historic')

tempList = [[ cell.value for cell in Header ] for Header in wsPast['B3':'G3']]
Anlagentypen = tempList[0] #unschön wegen unkenntnis
       
RE_cap = {}
for data in range(4, wsPast.max_row -2): # Spalten
    intYear = wsPast[get_column_letter(1)+str(data)].value
    for col in range(2, wsPast.max_column+1): 
        tempType = Anlagentypen[col-2] # wegen Liste und Excel 0 und 1
        RE_cap.setdefault(tempType,{})
        RE_cap[tempType].setdefault(intYear, {'cap_total':0, 'cap_year':0})                   
        RE_cap[tempType][intYear]['cap_total'] = wsPast.cell(row=data, column=col).value

#Add Zubaupfad
zubau = wb.get_sheet_by_name('zubau')    
for Anlagentyp in range(2, zubau.max_row+1):
    for col in range(2, zubau.max_column+1): #2 wegen ab 2.Spalte in Excel
        intYear = zubau[get_column_letter(col)+'1'].value #1. Zeile Header = Jahreszahlen
        strAnlage = zubau.cell(row=Anlagentyp, column=1).value
        RE_cap[strAnlage].setdefault(intYear, {'cap_total':0, 'cap_year':0})
        RE_cap[strAnlage][intYear]['cap_year']= zubau.cell(row=Anlagentyp, column=col).value
        
#Erzeuge fehlende Daten
#2030+ 
for k  in RE_cap:
    for i in range(iStart,iEnd+1):
        if i == iStart:
            RE_cap[k][i]['cap_year'] = RE_cap[k][i]['cap_total']
        else:
            try:
                if RE_cap[k][i]['cap_total']>0 :
                    RE_cap[k][i]['cap_year'] = RE_cap[k][i]['cap_total']-RE_cap[k][i-1]['cap_total']
                    if RE_cap[k][i]['cap_year'] < 0:
                        RE_cap[k][i]['cap_year'] = 0 # Bei Wasser der Fall oder wenn es wirklich zu Nettoabnahme kommt
                        # if RE_cap[k][i]['cap_total']-RE_cap[k][i-1]['cap_total'] <0:
                        # print ('Fehler in historischen Werten bei' + k + str(i))
                else:
                    try:
                        RE_cap[k][i]['cap_total'] = RE_cap[k][i-1]['cap_total'] + RE_cap[k][i]['cap_year']
                    except TypeError: # der Fall, wenn kein Zubau angegeben
                        RE_cap[k][i]['cap_total'] = RE_cap[k][i-1]['cap_total']
                        RE_cap[k][i]['cap_year'] = 0
            except KeyError: 
                RE_cap[k].setdefault(i, {'cap_total':0, 'cap_year':0}) 
                RE_cap[k][i]['cap_year'] = RE_cap[k][i-1]['cap_year'] # Forschreibung mit letztem Ausbau (Annahme!)
                RE_cap[k][i]['cap_total'] = RE_cap[k][i-1]['cap_total'] + RE_cap[k][i]['cap_year']
print ('Initial data received.')                
##################################
##Datenaufnahme abgeschlossen#####
##################################

#listDecay = readDecays()
#listEcoLife = readEcoLifes()
#setCapMatrix = [[capMatrix(Anlagentypen, listDecay[i], listEcoLifes[0])] for i in range(len(listDecay))]
setCapMatrix = []
for i in range (len(readDecays())):
    setCapMatrix.append(capMatrix(Anlagentypen, readDecays()[i], iStart, iEnd, readEcoLifes()[0], RE_cap))

listResultCap  = []
for i in range (len(setCapMatrix)):
    listResultCap.append (sumCapMatrix(Anlagentypen,setCapMatrix[i]))

listResultPow  = {}
for i in range (len(listResultCap)):
    strDecay = readDecaysAsName(i)
    listResultPow.setdefault(strDecay, [])
    listResultPow[strDecay]= (sumPowerMatrix(listResultCap[i], iStart, iEnd, Anlagentypen))
print('Berechnungen abgeschlossen. Starte Exportierung nach Excel...')

writer = pd.ExcelWriter('Backup.xlsx', engine='openpyxl')
writer.book = load_workbook(strWorkbook)
df1 = pd.DataFrame(listResultPow)
df1.to_excel(writer,sheet_name= 'result_Pow')

df2 = pd.DataFrame(listResultCap)

i=0
for cases in range(len(Anlagentypen)):    
    pd.DataFrame(df2[Anlagentypen[cases]]).to_excel(writer, sheet_name ='result_Cap_df2', startrow=i) #!!!
    #df2[Anlagentypen[cases]].to_frame(3).to_excel(writer, sheet_name ='result_Cap', startrow=i)
    i = i + len(readDecays()) +2
   # df2['Bio'].to_excel(writer,sheet_name= 'results2', startrow=i, startcol=1)
writer.save()
'''
#TWh für 2006 bis 2012 laut BMWi
RefPow = [ 47.7  ,	62.6  ,	68.1  ,	71.5  	,79.1  ,	101.3  	,116.7  , 123.9, 136.8]
print ('Nach BMWi ab 2006: ', RefPow , ' +25 TWh Wasserkraft')           
'''
print('Daten ausgeschrieben, erzeuge Graphiken...')
#Graphik zu Kapazitäten
#...
#Graphik zu Power
'''
from openpyxl.charts import Reference, Series, LineChart 
wb=openpyxl.load_workbook('Backup.xlsx')
ws_resPow = wb.get_sheet_by_name('result_Pow')

chartPow = LineChart()
chartPow.drawing.name = 'This is my first chart'
values = Reference(wb, (1,1), (2,60))
series = Series(values, title ="1. Run")
ws_resPow.add_chart(chartPow)
wb.save("Backup_print.xslx")
print('Graphiken erzeugt. Daten gespeichert. Ende')
'''